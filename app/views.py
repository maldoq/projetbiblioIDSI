import csv
from datetime import datetime
import openpyxl

from django.http import JsonResponse
from django.utils import timezone
from django.utils.timezone import now, timedelta
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.utils.text import slugify
from django.db.models import Q, F, Count
from django.db import transaction
from django.core.paginator import Paginator
from django.http import HttpResponse

from books.models import ActivityLog, Categorie, Emprunter, Etudiant, Ecole, Editeur, Auteur, Livre,ICONE_CHOICES,LANGUAGES_CHOICES,ETAT_LIVRE_CHOICES

# Authentification

def signin(request):
    # Si l’utilisateur est déjà connecté → redirection
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        remember = request.POST.get("remember")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            # Gestion du "remember me"
            if not remember:
                request.session.set_expiry(0)  # expire quand on ferme le navigateur

            messages.success(request, "Connexion réussie !")
            return redirect("dashboard")
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect.")
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    messages.info(request, "Vous avez été déconnecté.")
    return redirect('signin')



# Dashboard
@login_required(login_url='signin')
def dash(request):
    today = now().date()
    last_week = today - timedelta(days=6)  # pour les 7 derniers jours

    # --- KPI Cards ---
    total_books = sum([book.available_quantity() for book in Livre.objects.all()])
    active_loans = Emprunter.objects.filter(status="active").count()
    late_returns = Emprunter.objects.filter(status="late").count()
    total_users = Etudiant.objects.filter(is_active=True).count()

    late_returns_with_returned = Emprunter.objects.filter(
    Q(dateRetourEffectif__gt=F('dateRetourPrevu')) | 
    Q(dateRetourPrevu__lt=now().date(), dateRetourEffectif__isnull=True),
    status="returned"
).count()

    # --- Graphiques ---
    # 1) Emprunts par jour (line chart)
    loans_last_7_days = []
    labels_last_7_days = []
    for i in range(7):
        day = last_week + timedelta(days=i)
        labels_last_7_days.append(day.strftime("%a"))  # Lun, Mar, ...
        loans_count = Emprunter.objects.filter(dateEmprunt=day).count()
        loans_last_7_days.append(loans_count)

    # 2) Disponibilité des livres (donut chart)
    total_disponibles = sum([book.available_quantity() for book in Livre.objects.all()])
    total_empruntes = sum([book.quantite - book.available_quantity() for book in Livre.objects.all()])
    total_en_retard = sum([1 for emprunt in Emprunter.objects.all() if emprunt.is_overdue()])

    # 3) Statut des emprunts (bar chart)
    status_counts = Emprunter.objects.values('status').annotate(count=Count('id'))
    status_dict = {s['status']: s['count'] for s in status_counts}
    loans_status = {
        "late": status_dict.get("late", 0),
        "returned": status_dict.get("returned", 0),
        "active": status_dict.get("active", 0)
    }

    # --- Activités récentes ---
    recent_activities = Emprunter.objects.all().order_by("-dateEmprunt")[:3]
    activities_placeholder = recent_activities if recent_activities else []

    context = {
        "total_books": total_books,
        "active_loans": active_loans,
        "late_returns": late_returns_with_returned,
        "total_users": total_users,
        "labels_last_7_days": labels_last_7_days,
        "loans_last_7_days": loans_last_7_days,
        "total_disponibles": total_disponibles,
        "total_empruntes": total_empruntes,
        "total_en_retard": total_en_retard,
        "loans_status": loans_status,
        "recent_activities": recent_activities,
        "activities_placeholder": activities_placeholder
    }
    return render(request, 'dashboard.html', context)



# Gestion des livres

@login_required(login_url='signin')
def books_list(request):
    categories_active = Categorie.objects.filter(is_active=True)
    livres = Livre.objects.all()

    search = request.GET.get("search", "")
    category_id = request.GET.get("category", "")
    status = request.GET.get("status", "")

    if search:
        livres = livres.filter(
            Q(titre__icontains=search) |
            Q(isbn__icontains=search) |
            Q(auteur__nom_complet__icontains=search) |
            Q(editeur__nom__icontains=search)
        )

    if category_id:
        livres = livres.filter(categorie_id=category_id)

    if status == "available":
        livres = livres.filter(quantite__gte=1)

    elif status == "borrowed":
        livres = livres.filter(quantite__lt=1)

    paginator = Paginator(livres, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "books_list.html", {
        "books": page_obj,
        "categories": categories_active,
        "search_query": search,
        "selected_category": category_id,
        "selected_status": status,
        "page_obj": page_obj,
        "is_paginated": True,
    })

@login_required(login_url="signin")
def books_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livres"

    headers = [
        "isbn",
        "titre",
        "langue",
        "quantite",
        "nbre_pages",
        "annee_publication",
        "emplacement",
        "resume",
        "editeur",
        "auteur",
        "categorie",
    ]
    ws.append(headers)

    livres = Livre.objects.select_related(
        "editeur", "auteur", "categorie"
    ).all()

    for livre in livres:
        ws.append([
            livre.isbn,
            livre.titre,
            livre.langue,
            livre.quantite,
            livre.nbre_pages,
            livre.annee_publication or "",
            livre.emplacement or "",
            livre.resume or "",
            livre.editeur.nom,
            livre.auteur.nom_complet,
            livre.categorie.nom,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="livres.xlsx"'
    wb.save(response)

    return response

@login_required(login_url="signin")
def books_import_excel(request):
    if request.method != "POST":
        return redirect("books_list")

    excel_file = request.FILES.get("excel_file")
    if not excel_file:
        messages.error(request, "Aucun fichier sélectionné.")
        return redirect("books_list")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    headers = [h.strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    REQUIRED_COLS = {
        "isbn",
        "titre",
        "langue",
        "quantite",
        "nbre_pages",
        "editeur",
        "auteur",
        "categorie",
    }

    missing = REQUIRED_COLS - set(headers)
    if missing:
        messages.error(
            request,
            f"Colonnes manquantes : {', '.join(missing)}"
        )
        return redirect("books_list")

    col = {name: headers.index(name) for name in headers}

    try:
        with transaction.atomic():
            for line_no, row in enumerate(data_rows, start=2):
                try:
                    editeur, _ = Editeur.objects.get_or_create(
                        nom=row[col["editeur"]]
                    )
                    auteur, _ = Auteur.objects.get_or_create(
                        nom_complet=row[col["auteur"]]
                    )
                    categorie, _ = Categorie.objects.get_or_create(
                        nom=row[col["categorie"]],
                        defaults={"is_active": True}
                    )

                    Livre.objects.update_or_create(
                        isbn=row[col["isbn"]],
                        defaults={
                            "titre": row[col["titre"]],
                            "langue": row[col["langue"]],
                            "quantite": int(row[col["quantite"]]),
                            "nbre_pages": int(row[col["nbre_pages"]]),
                            "annee_publication": row[col.get("annee_publication")] or None,
                            "emplacement": row[col.get("emplacement")] or "",
                            "resume": row[col.get("resume")] or "",
                            "editeur": editeur,
                            "auteur": auteur,
                            "categorie": categorie,
                        }
                    )

                except Exception as e:
                    raise ValueError(f"Ligne {line_no} : {e}")

        messages.success(request, "Importation des livres réussie.")

    except Exception as e:
        messages.error(request, f"Import annulé : {e}")

    return redirect("books_list")

@login_required(login_url='signin')
def search_authors(request):
    """
    Endpoint API pour rechercher des auteurs
    URL: /api/authors/search/?q=<query>
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse([], safe=False)
    
    # Recherche insensible à la casse
    authors = Auteur.objects.filter(
        nom_complet__icontains=query
    ).order_by('nom_complet')[:10]  # Limiter à 10 résultats
    
    # Formatter les résultats
    results = []
    for author in authors:
        results.append({
            'id': author.id,
            'nom_complet': author.nom_complet,
            'dateNaiss': author.dateNaiss.isoformat() if author.dateNaiss else None
        })
    
    return JsonResponse(results, safe=False)


@login_required(login_url='signin')
def search_publishers(request):
    """
    Endpoint API pour rechercher des éditeurs
    URL: /api/publishers/search/?q=<query>
    """
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse([], safe=False)
    
    # Recherche insensible à la casse
    publishers = Editeur.objects.filter(
        nom__icontains=query
    ).order_by('nom')[:10]  # Limiter à 10 résultats
    
    # Formatter les résultats
    results = []
    for publisher in publishers:
        results.append({
            'id': publisher.id,
            'nom': publisher.nom
        })
    
    return JsonResponse(results, safe=False)

@login_required(login_url='signin')
def books_form(request, pk=None):
    """
    Formulaire d'ajout/modification de livre avec autocomplétion
    """
    # Mode modification si un ID est présent
    book = None
    if pk:
        book = get_object_or_404(Livre, pk=pk)
    
    if request.method == "POST":
        isbn = request.POST.get("isbn")
        titre = request.POST.get("title")
        langue = request.POST.get("language")
        quantite = request.POST.get("total_quantity")
        nbre_pages = request.POST.get("pages")
        annee_publication = request.POST.get("publication_year")
        emplacement = request.POST.get("location")
        resume = request.POST.get("description")
        
        # Récupération des informations auteur
        author_id = request.POST.get("author_id")
        author_name = request.POST.get("author")
        
        # Récupération des informations éditeur
        publisher_id = request.POST.get("publisher_id")
        publisher_name = request.POST.get("publisher")
        
        categorie_id = request.POST.get("category")

        # Gestion de l'éditeur
        if publisher_id:
            # Éditeur existant sélectionné
            editeur_obj = get_object_or_404(Editeur, id=publisher_id)
        else:
            # Création d'un nouvel éditeur
            editeur_obj, created = Editeur.objects.get_or_create(
                nom=publisher_name.upper()
            )

        # Gestion de l'auteur
        if author_id:
            # Auteur existant sélectionné
            auteur_obj = get_object_or_404(Auteur, id=author_id)
        else:
            # Création d'un nouvel auteur
            auteur_obj, created = Auteur.objects.get_or_create(
                nom_complet=author_name.title()
            )

        categorie = get_object_or_404(Categorie, id=categorie_id)

        if book:
            # Modification
            book.isbn = isbn
            book.titre = titre
            book.langue = langue
            book.quantite = quantite
            book.nbre_pages = nbre_pages if nbre_pages else None
            book.annee_publication = annee_publication if annee_publication else None
            book.emplacement = emplacement
            book.resume = resume
            book.editeur = editeur_obj
            book.auteur = auteur_obj
            book.categorie = categorie
            book.save()

            ActivityLog.objects.create(
                action_type="update",
                title=f"Modification du livre {book.titre}",
                description=f"« {book.titre} » ({book.quantite} exemplaires)",
                performed_by=request.user
            )
        else:
            # Création
            livre = Livre.objects.create(
                isbn=isbn,
                titre=titre,
                langue=langue,
                quantite=quantite,
                nbre_pages=nbre_pages if nbre_pages else None,
                annee_publication=annee_publication if annee_publication else None,
                emplacement=emplacement,
                resume=resume,
                editeur=editeur_obj,
                auteur=auteur_obj,
                categorie=categorie
            )

            ActivityLog.objects.create(
                action_type="add_book",
                title="Nouveau livre ajouté",
                description=f"« {livre.titre} » ({livre.quantite} exemplaires)",
                performed_by=request.user
            )

        return redirect("books_list")

    categories_active = Categorie.objects.filter(is_active=True)
    return render(request, 'books_form.html', {
        'book': book,
        'categories': categories_active,
        'languages': LANGUAGES_CHOICES
    })

@login_required(login_url='signin')
def books_delete(request,pk):
    book = get_object_or_404(Livre, pk=pk)
    ActivityLog.objects.create(
        action_type="delete",
        title=f"Suppression du livre {book.titre}",
        description=f"« {book.titre} » ({book.quantite} exemplaires)",
        performed_by=request.user
    )
    book.delete()
    return redirect("books_list")



# Gestion des catégories

@login_required(login_url='signin')
def categories_list(request):
    categories = Categorie.objects.all()
    return render(request, "categories_list.html", {
        "categories": categories
    })

@login_required(login_url='signin')
def categories_form(request, pk=None):
    # Mode modification si un ID est présent
    category = None
    if pk:
        category = get_object_or_404(Categorie, pk=pk)

    if request.method == "POST":
        nom = request.POST.get("nom")
        description = request.POST.get("description")
        icone = request.POST.get("icone")
        couleur = request.POST.get("couleur")
        active = request.POST.get("is_active") is not None

        if category:
            # Modification
            category.nom = nom
            category.description = description
            category.icone = icone
            category.couleur = couleur
            category.slug_url = slugify(nom)
            category.is_active = active
            category.save()

            ActivityLog.objects.create(
                action_type="update",
                title=f"Modification de la catégorie {category.nom}",
                description=f"« {category.nom} » modifiée",
                performed_by=request.user
            )
        else:
            # Création
            category = Categorie.objects.create(
                nom=nom,
                description=description,
                icone=icone,
                couleur=couleur,
                is_active=active,
                slug_url=slugify(nom)
            )

            ActivityLog.objects.create(
                action_type="add_book",
                title="Nouvelle catégorie ajoutée",
                description=f"« {category.nom} » ajoutée",
                performed_by=request.user
            )

        return redirect("categories_list")

    return render(request, "categories_form.html", {
        "category": category,
        "ICONE_CHOICES": ICONE_CHOICES,
    })

@login_required(login_url='signin')
def categories_delete(request,pk):
    category = get_object_or_404(Categorie, pk=pk)
    ActivityLog.objects.create(
        action_type="delete",
        title=f"Suppression de la catégorie {category.nom}",
        description=f"« {category.nom} » supprimé",
        performed_by=request.user
    )
    category.delete()
    return redirect("categories_list")



# Gestion des Etudiants

@login_required(login_url='signin')
def users_list(request):
    users = Etudiant.objects.all()

    search = request.GET.get("search", "")

    if search:
        users = users.filter(
                Q(matricule__icontains=search) |
                Q(nom__icontains=search) |
                Q(emailInst__icontains=search) |
                Q(telephone__icontains=search)
            )
        
    paginator = Paginator(users, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, 'users_list.html', {
        "users": page_obj,
        "search_query": search,
        "page_obj": page_obj,
        "is_paginated": True,
    })

@login_required(login_url='signin')
def users_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    # ⚠️ Les noms DOIVENT correspondre à l'import
    headers = [
        "matricule",
        "nom",
        "prenoms",
        "email",
        "telephone",
        "ecole",
        "dateNaiss",
        "actif",
    ]
    ws.append(headers)

    users = Etudiant.objects.select_related("ecole").all()

    for u in users:
        ws.append([
            u.matricule,
            u.nom,
            u.prenoms,
            u.emailInst,
            u.telephone,
            u.ecole.nom if u.ecole else "",
            u.dateNaiss.strftime("%Y-%m-%d") if u.dateNaiss else "",
            "Oui" if u.is_active else "Non",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="utilisateurs.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='signin')
def users_import_excel(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Aucun fichier sélectionné.")
            return redirect("users_list")

        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        data_rows = rows[1:]

        for index, row in enumerate(data_rows, start=2):
            try:
                (
                    matricule,
                    nom,
                    prenoms,
                    email,
                    telephone,
                    ecole_nom,
                    date_naiss,
                    actif,
                ) = row

                if not date_naiss:
                    raise ValueError(f"Date de naissance manquante (ligne {index})")

                # Conversion date si besoin
                if isinstance(date_naiss, str):
                    date_naiss = datetime.strptime(date_naiss, "%Y-%m-%d").date()

                ecole = None
                if ecole_nom:
                    ecole, _ = Ecole.objects.get_or_create(nom=ecole_nom)

                Etudiant.objects.update_or_create(
                    matricule=matricule,
                    defaults={
                        "nom": nom,
                        "prenoms": prenoms,
                        "emailInst": email,
                        "telephone": telephone,
                        "ecole": ecole,
                        "dateNaiss": date_naiss,
                        "is_active": True if actif == "Oui" else False,
                    }
                )

            except Exception as e:
                messages.error(
                    request,
                    f"Erreur ligne {index} : {e}"
                )
                return redirect("users_list")

        messages.success(request, "Importation Excel réussie.")

    return redirect("users_list")

@login_required(login_url='signin')
def users_form(request, pk=None):

    # Mode modification si un ID est présent
    user = None
    if pk:
        user = get_object_or_404(Etudiant, pk=pk)

    if request.method == "POST":
        matricule = request.POST.get("matricule")
        nom = request.POST.get("nom")
        prenoms = request.POST.get("prenoms")
        telephone = request.POST.get("telephone")
        emailInst = request.POST.get("emailInst")
        emailPers = request.POST.get("emailPers")
        ecole_nom = request.POST.get("ecole")
        dateNaiss = request.POST.get("dateNaiss")
        numChambre = request.POST.get("numChambre")
        is_active = request.POST.get("is_active") is not None

        # Sélection ou création d'une école
        ecole_obj, created = Ecole.objects.get_or_create(
            nom=ecole_nom.upper()
        )

        if user:
            # Modification
            user.matricule = matricule
            user.nom = nom
            user.prenoms = prenoms
            user.telephone = telephone
            user.emailInst = emailInst
            user.emailPers = emailPers
            user.dateNaiss = dateNaiss
            user.numChambre = numChambre
            user.ecole = ecole_obj
            user.is_active = is_active
            user.save()
            ActivityLog.objects.create(
                action_type="update",
                title=f"Modification de l'etudiant {user.nom}",
                description=f"« {user.nom} » modifiée",
                performed_by=request.user
            )
        else:
            # Création
            etudiant = Etudiant.objects.create(
                matricule = matricule,
                nom = nom,
                prenoms = prenoms,
                dateNaiss = dateNaiss,
                telephone = telephone,
                emailInst = emailInst,
                emailPers = emailPers,
                numChambre = numChambre,
                ecole = ecole_obj,
                is_active = is_active

            )

            ActivityLog.objects.create(
                action_type="add_user",
                title="Nouvel étudiant ajouté",
                description=f"« {etudiant.nom} {etudiant.prenoms} » ajouté",
                performed_by=request.user
            )
        return redirect("users_list")


    return render(request, 'users_form.html',{
        "user":user
    })

@login_required(login_url='signin')
def users_delete(request,pk):
    user = Etudiant.objects.get(pk=pk)
    ActivityLog.objects.create(
        action_type="delete",
        title=f"Suppression de l'utilisateur {user.nom} {user.prenoms}",
        description=f"« {user.nom} {user.prenoms} » supprimé",
        performed_by=request.user
    )
    user.delete()
    return redirect("users_list")



# Gestion des emprunts

@login_required(login_url='signin')
def loans_list(request):
    emprunts = Emprunter.objects.select_related(
        "etudiant",
        "livre",
        "livre__auteur"
    )

    search = request.GET.get("search", "")
    status = request.GET.get("status", "all")

    # 🔍 SEARCH
    if search:
        emprunts = emprunts.filter(
            Q(etudiant__nom__icontains=search) |
            Q(livre__titre__icontains=search) |
            Q(livre__auteur__nom_complet__icontains=search)
        )

    today = timezone.now().date()

    # 🎯 FILTER STATUS
    if status == "active":
        emprunts = emprunts.filter(
            dateRetourEffectif__isnull=True,
            dateRetourPrevu__gte=today
        )

    elif status == "late":
        emprunts = emprunts.filter(
            dateRetourEffectif__isnull=True,
            dateRetourPrevu__lt=today
        )

    elif status == "returned":
        emprunts = emprunts.filter(
            dateRetourEffectif__isnull=False
        )

    # 📄 PAGINATION
    paginator = Paginator(emprunts.order_by("-dateEmprunt"), 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "loans_list.html", {
        "loans": page_obj,
        "search_query": search,
        "status_filter": status,
        "page_obj": page_obj,
        "is_paginated": True,
    })

@login_required(login_url='signin')
def loans_form(request, pk=None):
    # Mode de visualisation
    loan = None
    if pk:
        loan = get_object_or_404(Emprunter,pk=pk)

    if request.method == "POST":
        etudiant_matri = request.POST.get("user")
        livre_isbn = request.POST.get("book")
        dateEmprunt = request.POST.get("borrow_date")
        dateRetourPrevu = request.POST.get("due_date")
        observation = request.POST.get("notes")

        etudiant = get_object_or_404(Etudiant, matricule=etudiant_matri)
        livre = get_object_or_404(Livre, isbn=livre_isbn)

        if not loan:
            # Création
            Emprunter.objects.create(
                etudiant = etudiant,
                livre = livre,
                dateEmprunt = dateEmprunt,
                dateRetourPrevu = dateRetourPrevu,
                observation = observation,
                status = "active"
            )

            ActivityLog.objects.create(
                action_type="loan",
                title="Nouvel emprunt",
                description=f"{etudiant.nom} {etudiant.prenoms} a emprunté « {livre.titre} »",
                user=str(etudiant.nom + ' ' + etudiant.prenoms),
                performed_by=request.user
            )

        return redirect("loans_list")

    context = {
        "loan":loan,
        'today': timezone.now().date(),
        'available_books': Livre.disponibles(),
        "users": Etudiant.objects.filter(is_active=True)
    }
    return render(request, 'loans_form.html', context)

@login_required(login_url='signin')
def returns_form(request):

    if request.method == "POST":
        loan_id = request.POST.get("loan")
        dateRetour = request.POST.get("return_date")
        etat = request.POST.get("condition")
        notes = request.POST.get("notes")

        loan = get_object_or_404(Emprunter,id=loan_id)

        loan.dateRetourEffectif = dateRetour
        loan.etat_livre = etat
        loan.observation += " \n" + notes
        loan.status = "returned"
        loan.save()

        ActivityLog.objects.create(
            action_type="return",
            title="Retour de livre",
            description=f"{loan.etudiant.nom} a retourné « {loan.livre.titre} »",
            user=str(loan.etudiant),
            performed_by=request.user
        )

        return redirect("loans_list")

    emprunts = Emprunter.objects.filter(status__in=["late","active"])
    context = {
        "active_loans":emprunts,
        "etats": ETAT_LIVRE_CHOICES,
    }
    return render(request, 'returns_form.html',context)

@login_required(login_url='signin')
def loans_delete(request, pk):
    loan = get_object_or_404(Emprunter, pk=pk)
    ActivityLog.objects.create(
        action_type="delete",
        title=f"Suppression de l'emprunt de {loan.etudiant.nom} {loan.etudiant.prenoms} du livre {loan.livre.titre}",
        description=f"« emprunt numero {loan.id} » supprimé",
        performed_by=request.user
    )
    loan.delete()
    return redirect("loans_list")



# Gestion du profil des utilisateur

@login_required(login_url="signin")
def history(request):
    activities = ActivityLog.objects.all().order_by("-timestamp")

    search = request.GET.get("search", "")
    action_type = request.GET.get("action_type", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")

    if search:
        activities = activities.filter(
            Q(title__icontains=search) |
            Q(description__icontains=search) |
            Q(user__icontains=search)
        )

    if action_type:
        activities = activities.filter(action_type=action_type)

    if date_from:
        activities = activities.filter(timestamp__date__gte=date_from)

    if date_to:
        activities = activities.filter(timestamp__date__lte=date_to)

    paginator = Paginator(activities, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "history.html", {
        "activities": page_obj,
        "search_query": search,
        "action_type": action_type,
        "date_from": date_from,
        "date_to": date_to,
        "page_obj": page_obj,
        "is_paginated": True,
    })

@login_required(login_url='signin')
def profile(request):
    return render(request, 'profile.html')

@login_required(login_url="signin")
def history_export(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="historique.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "Date", "Action", "Titre", "Description", "Utilisateur", "Effectué par"
    ])

    for act in ActivityLog.objects.all().order_by("-timestamp"):
        writer.writerow([
            act.timestamp.strftime("%d/%m/%Y %H:%M"),
            act.action_type,
            act.title,
            act.description,
            act.user,
            act.performed_by.username if act.performed_by else ""
        ])

    return response

@login_required(login_url='signin')
def change_password(request):
    pass

# Sécurité
def custom_404(request, exception):
    return render(request,"404.html",status=404)
